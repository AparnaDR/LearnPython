# program to load excel and reduce the price by 10% on 4th row
# step 2 is to create a bar chart
import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    """
    cell = sheet['a1'] # we can get the value using the excel values
    cell = sheet.cell(1,1) # or through the index
    print(cell.value) # prints the value
    print(sheet.max_row) """

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)  # create row 4 for putting the output
        corrected_price_cell.value = corrected_price

    # add chart
    # select range of values. Use Reference class
    # select all the cell from the 4th col from row 2 to maximum last row
    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    # now we need to create a bar chart
    # create an instance if bar chart class

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)
